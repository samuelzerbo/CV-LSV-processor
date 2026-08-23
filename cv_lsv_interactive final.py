"""
cv_lsv_interactive.py
=======================
Unified interactive tool for processing Cyclic Voltammetry (CV) and Linear
Sweep Voltammetry (LSV) raw data. Pop-up windows walk you through:

  0. Choose whether this file is CV or LSV data
  1. Upload the raw data file (.csv or .xlsx)
     -> the file's columns are validated against what's required for the
        chosen type. If something's missing, you get a clear error telling
        you exactly what's wrong and a chance to pick a different file --
        it will NOT silently process a wrong/mismatched file.
  2. RHE conversion formula: pick a reference electrode + enter pH
  3. Electrode surface area (cm^2)
  4. Save-file dialog for the output Excel

After each file finishes, it asks "process another file?" -- say yes to
loop back to step 0 (you can mix CV and LSV files in the same session).

STRUCTURE REQUIREMENTS
-------------------------
CV  needs columns: Scan, Index, WE(1).Potential (V), WE(1).Current (A)
    -- and the Scan column must contain scan numbers 2 through 10.
LSV needs columns: Index, WE(1).Potential (V), WE(1).Current (A)

WHAT IT COMPUTES
------------------
CV:  averages scan 2-10 current (aligned by position within each scan),
     builds: Potential | 2 | 3 | ... | 10 | avarage | Potential RHE | current density mA/cm2
LSV: converts every row directly:
     Potential RHE = WE(1).Potential (V) + offset + 0.0591 * pH
     Current density mA/cm2 = (WE(1).Current (A) * 1000) / area

Both write "Potential RHE" / "current density" (and CV's "avarage") as LIVE
EXCEL FORMULAS linked to editable yellow input cells (columns AA:AB) -- edit
pH / offset / area there later and everything recalculates automatically.

REQUIREMENTS
------------
Needs a normal desktop Python install with tkinter (ships by default with
Python on Windows/Mac; on Linux: `sudo apt install python3-tk`).
    pip install pandas openpyxl
"""

import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox

SCAN_COL = "Scan"
INDEX_COL = "Index"
POTENTIAL_COL = "WE(1).Potential (V)"
CURRENT_COL = "WE(1).Current (A)"
REQUIRED_CV_SCANS = list(range(2, 11))  # 2..10

RAW_COLUMNS_ORDER_CV = [
    "Potential applied (V)", "Time (s)", "WE(1).Current (A)", "WE(1).Potential (V)",
    "Scan", "Index", "Q+", "Q-", "Current range",
]
RAW_COLUMNS_ORDER_LSV = [
    "Potential applied (V)", "Time (s)", "WE(1).Current (A)", "WE(1).Potential (V)",
    "Index", "Current range",
]

FONT = "Arial"
INPUT_FONT = Font(name=FONT, color="0000FF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
LABEL_FONT = Font(name=FONT, bold=True)
HEADER_FONT = Font(name=FONT, bold=True)
BODY_FONT = Font(name=FONT)

REF_ELECTRODES = [
    ("Ag/AgCl, saturated KCl", 0.197),
    ("Ag/AgCl, 3 M KCl", 0.210),
    ("Ag/AgCl, 3 M saturated KCl", 0.205),
    ("Ag/AgCl, 3 M NaCl", 0.209),
    ("Ag/AgCl, 1 M KCl", 0.235),
    ("SCE (saturated calomel)", 0.241),
    ("Hg/HgO, 1 M NaOH", 0.140),
    ("Custom offset...", None),
]


# --------------------------- loading & validation ---------------------------

def load_raw_data(path: str) -> pd.DataFrame:
    if path.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path, sep=None, engine="python")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def validate_structure(df: pd.DataFrame, kind: str):
    """Returns (ok: bool, message: str or None).

    Checks both required column names AND a structural signature that
    distinguishes CV from LSV: CV data has a 'Scan' column with MULTIPLE
    scan values (repeated cycles); LSV is a single sweep, so a 'Scan'
    column -- if present at all -- should only have one value. This catches
    the case where someone selects the wrong type for an otherwise
    column-compatible file (e.g. picking 'LSV' for an actual CV file).
    """
    if kind == "CV":
        required = [SCAN_COL, INDEX_COL, POTENTIAL_COL, CURRENT_COL]
    else:
        required = [INDEX_COL, POTENTIAL_COL, CURRENT_COL]

    missing = [c for c in required if c not in df.columns]
    if missing:
        return False, (
            f"This file is missing the following required column(s) for {kind} data:\n\n"
            f"  {', '.join(missing)}\n\n"
            f"Found columns:\n  {', '.join(str(c) for c in df.columns)}\n\n"
            f"Please select the correct raw {kind} data file."
        )

    if kind == "CV":
        scans_present = set(pd.to_numeric(df[SCAN_COL], errors="coerce").dropna().astype(int).unique().tolist())
        missing_scans = [s for s in REQUIRED_CV_SCANS if s not in scans_present]
        if missing_scans:
            return False, (
                f"This file's '{SCAN_COL}' column doesn't include all of scans 2 through 10.\n\n"
                f"Missing scan(s): {missing_scans}\n"
                f"Scans found in file: {sorted(scans_present)}\n\n"
                f"Please select the correct raw CV data file (with scans 2-10)."
            )

    if kind == "LSV" and SCAN_COL in df.columns:
        scans_present = set(pd.to_numeric(df[SCAN_COL], errors="coerce").dropna().astype(int).unique().tolist())
        if len(scans_present) > 1:
            return False, (
                f"This file has a '{SCAN_COL}' column with {len(scans_present)} different scan values "
                f"({sorted(scans_present)}).\n\n"
                f"That's the signature of CV data (repeated cycles), not LSV (a single sweep).\n\n"
                f"Please choose 'CV' instead, or select the correct raw LSV data file."
            )

    return True, None


# --------------------------- CV processing ---------------------------

def build_cv_wide_table(df: pd.DataFrame, min_scan: int, max_scan: int, potential_from_scan: int):
    sub = df[df[SCAN_COL].between(min_scan, max_scan)].copy()
    sub = sub.sort_values([SCAN_COL, INDEX_COL])
    sub["_point"] = sub.groupby(SCAN_COL).cumcount()

    pivot = sub.pivot_table(index="_point", columns=SCAN_COL, values=CURRENT_COL)
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)

    pot_sub = sub[sub[SCAN_COL] == potential_from_scan][["_point", POTENTIAL_COL]]
    pot_sub = pot_sub.set_index("_point").rename(columns={POTENTIAL_COL: "Potential"})

    out = pot_sub.join(pivot, how="inner").reset_index(drop=True)
    scan_numbers = list(pivot.columns)
    return out, scan_numbers


def write_cv_workbook(raw_df, table_df, scan_numbers, ph, ref_offset_v, area_cm2, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "CV Data"

    raw_cols = [c for c in RAW_COLUMNS_ORDER_CV if c in raw_df.columns] or list(raw_df.columns)
    for j, col in enumerate(raw_cols, start=1):
        ws.cell(row=1, column=j, value=col).font = HEADER_FONT
    for i, row in enumerate(raw_df[raw_cols].itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val).font = BODY_FONT

    in_col_label, in_col_value = _write_inputs_block(ws, ph, ref_offset_v, area_cm2)

    start_col = 13
    headers = ["Potential"] + [str(s) for s in scan_numbers] + ["avarage", "Potential RHE", "current density mA/cm2"]
    for j, h in enumerate(headers, start=start_col):
        ws.cell(row=1, column=j, value=h).font = HEADER_FONT

    scan_col_letters = [get_column_letter(start_col + 1 + k) for k in range(len(scan_numbers))]
    pot_col_letter = get_column_letter(start_col)
    avg_col_letter = get_column_letter(start_col + 1 + len(scan_numbers))
    area_addr = f"${get_column_letter(in_col_value)}$4"
    combined_offset_addr = f"${get_column_letter(in_col_value)}$5"

    for i in range(len(table_df)):
        r = i + 2
        ws.cell(row=r, column=start_col, value=float(table_df["Potential"].iloc[i])).font = BODY_FONT
        for k, s in enumerate(scan_numbers):
            ws.cell(row=r, column=start_col + 1 + k, value=float(table_df[s].iloc[i])).font = BODY_FONT

        first_scan_col, last_scan_col = scan_col_letters[0], scan_col_letters[-1]
        ws.cell(row=r, column=start_col + 1 + len(scan_numbers),
                value=f"=AVERAGE({first_scan_col}{r}:{last_scan_col}{r})").font = BODY_FONT
        ws.cell(row=r, column=start_col + 2 + len(scan_numbers),
                value=f"={pot_col_letter}{r}+{combined_offset_addr}").font = BODY_FONT
        ws.cell(row=r, column=start_col + 3 + len(scan_numbers),
                value=f"=({avg_col_letter}{r}*1000)/{area_addr}").font = BODY_FONT

    for j in range(1, 30):
        ws.column_dimensions[get_column_letter(j)].width = 14
    wb.save(out_path)


# --------------------------- LSV processing ---------------------------

def write_lsv_workbook(raw_df, ph, ref_offset_v, area_cm2, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "LSV Data"

    raw_cols = [c for c in RAW_COLUMNS_ORDER_LSV if c in raw_df.columns] or list(raw_df.columns)
    for j, col in enumerate(raw_cols, start=1):
        ws.cell(row=1, column=j, value=col).font = HEADER_FONT
    for i, row in enumerate(raw_df[raw_cols].itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val).font = BODY_FONT

    n_rows = len(raw_df)
    potential_col_letter = get_column_letter(raw_cols.index(POTENTIAL_COL) + 1)
    current_col_letter = get_column_letter(raw_cols.index(CURRENT_COL) + 1)

    in_col_label, in_col_value = _write_inputs_block(ws, ph, ref_offset_v, area_cm2)

    start_col = 13
    ws.cell(row=1, column=start_col, value="Potential RHE").font = HEADER_FONT
    ws.cell(row=1, column=start_col + 1, value="current density mA/cm2").font = HEADER_FONT

    area_addr = f"${get_column_letter(in_col_value)}$4"
    combined_offset_addr = f"${get_column_letter(in_col_value)}$5"

    for i in range(n_rows):
        r = i + 2
        ws.cell(row=r, column=start_col,
                value=f"={potential_col_letter}{r}+{combined_offset_addr}").font = BODY_FONT
        ws.cell(row=r, column=start_col + 1,
                value=f"=({current_col_letter}{r}*1000)/{area_addr}").font = BODY_FONT

    for j in range(1, 30):
        ws.column_dimensions[get_column_letter(j)].width = 14
    wb.save(out_path)


# --------------------------- shared helpers ---------------------------

def _write_inputs_block(ws, ph, ref_offset_v, area_cm2):
    in_col_label, in_col_value = 27, 28
    ws.cell(row=1, column=in_col_label, value="INPUTS (edit these)").font = LABEL_FONT

    ws.cell(row=2, column=in_col_label, value="pH").font = BODY_FONT
    c = ws.cell(row=2, column=in_col_value, value=ph); c.font, c.fill = INPUT_FONT, INPUT_FILL

    ws.cell(row=3, column=in_col_label, value="Reference electrode offset (V vs SHE)").font = BODY_FONT
    c = ws.cell(row=3, column=in_col_value, value=ref_offset_v); c.font, c.fill = INPUT_FONT, INPUT_FILL

    ws.cell(row=4, column=in_col_label, value="Electrode area (cm^2)").font = BODY_FONT
    c = ws.cell(row=4, column=in_col_value, value=area_cm2); c.font, c.fill = INPUT_FONT, INPUT_FILL

    ws.cell(row=5, column=in_col_label, value="Combined RHE offset (V) = offset + 0.0591*pH").font = BODY_FONT
    ws.cell(row=5, column=in_col_value,
            value=f"={get_column_letter(in_col_value)}3+0.0591*{get_column_letter(in_col_value)}2").font = BODY_FONT

    note = ws.cell(row=7, column=in_col_label,
                    value="pH, reference offset, and area are the values you entered when running the script. "
                          "Edit the yellow cells above to change them -- the table recalculates automatically.")
    note.font = Font(name=FONT, italic=True, size=9)
    ws.merge_cells(start_row=7, start_column=in_col_label, end_row=9, end_column=in_col_value + 3)
    return in_col_label, in_col_value


# --------------------------- GUI dialogs ---------------------------

class TypeChoiceDialog(simpledialog.Dialog):
    """Ask whether the file is CV or LSV data."""

    def body(self, master):
        tk.Label(master, text="What type of data is this file?", font=(FONT, 10, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 8))
        self.choice = tk.StringVar(value="CV")
        tk.Radiobutton(master, text="CV (Cyclic Voltammetry) -- multiple scans, averages scans 2-10",
                       variable=self.choice, value="CV").grid(row=1, column=0, sticky="w")
        tk.Radiobutton(master, text="LSV (Linear Sweep Voltammetry) -- single sweep",
                       variable=self.choice, value="LSV").grid(row=2, column=0, sticky="w")
        return None

    def apply(self):
        self.result = self.choice.get()


class RefElectrodeDialog(simpledialog.Dialog):
    def body(self, master):
        tk.Label(master, text="Select your reference electrode:", font=(FONT, 10, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))
        self.choice = tk.IntVar(value=0)
        for i, (name, offset) in enumerate(REF_ELECTRODES):
            label = name if offset is None else f"{name}  (offset = {offset} V)"
            tk.Radiobutton(master, text=label, variable=self.choice, value=i).grid(
                row=i + 1, column=0, columnspan=2, sticky="w")
        tk.Label(master, text="Custom offset (V, only if selected above):").grid(
            row=len(REF_ELECTRODES) + 1, column=0, sticky="w", pady=(8, 0))
        self.custom_entry = tk.Entry(master)
        self.custom_entry.grid(row=len(REF_ELECTRODES) + 1, column=1, pady=(8, 0))
        return None

    def apply(self):
        idx = self.choice.get()
        name, offset = REF_ELECTRODES[idx]
        if offset is None:
            try:
                offset = float(self.custom_entry.get())
            except ValueError:
                offset = 0.0
            name = "Custom"
        self.result = (name, offset)


def _bring_to_front(win):
    try:
        win.attributes("-topmost", True)
        win.lift()
        win.focus_force()
        win.after(200, lambda: win.attributes("-topmost", False))
    except tk.TclError:
        pass


def ask_for_valid_file(root, kind: str):
    """Repeatedly prompts for a file until a structurally valid one is chosen, or the user cancels."""
    while True:
        _bring_to_front(root)
        raw_path = filedialog.askopenfilename(
            title=f"Upload raw {kind} data file",
            filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv"), ("All files", "*.*")],
            parent=root,
        )
        if not raw_path:
            return None, None

        try:
            df = load_raw_data(raw_path)
        except Exception as e:
            messagebox.showerror("Error reading file", f"Could not read this file:\n\n{e}", parent=root)
            continue

        ok, msg = validate_structure(df, kind)
        if not ok:
            messagebox.showerror("File structure doesn't match", msg, parent=root)
            continue

        return df, raw_path


def process_one_file(root) -> bool:
    # ---- 0. CV or LSV? ----
    type_dlg = TypeChoiceDialog(root, title="Step 1 of 5: Data type")
    if type_dlg.result is None:
        return False
    kind = type_dlg.result

    # ---- 1. Upload + validate raw data ----
    _bring_to_front(root)
    messagebox.showinfo(f"{kind} Data Processor", f"Step 2 of 5: choose your raw {kind} data file.", parent=root)
    df, raw_path = ask_for_valid_file(root, kind)
    if df is None:
        messagebox.showwarning("Cancelled", "No valid file selected.", parent=root)
        return False

    # ---- 2. RHE formula ----
    dlg = RefElectrodeDialog(root, title="Step 3 of 5: Reference electrode")
    if dlg.result is None:
        return False
    ref_name, ref_offset = dlg.result

    _bring_to_front(root)
    ph = simpledialog.askfloat(
        "Step 3 of 5: pH", "Enter electrolyte pH:\n(E_RHE = E_measured + offset + 0.0591 x pH)",
        minvalue=0.0, maxvalue=14.0, parent=root,
    )
    if ph is None:
        return False

    # ---- 3. Surface area ----
    _bring_to_front(root)
    area = simpledialog.askfloat(
        "Step 4 of 5: Electrode area", "Enter electrode surface area (cm^2):", minvalue=0.000001, parent=root,
    )
    if area is None:
        return False

    # ---- 4. Calculate ----
    try:
        if kind == "CV":
            table_df, scan_numbers = build_cv_wide_table(df, min_scan=2, max_scan=10, potential_from_scan=2)
        else:
            table_df, scan_numbers = None, None
    except Exception as e:
        messagebox.showerror("Error processing data", str(e), parent=root)
        return False

    # ---- 5. Save ----
    _bring_to_front(root)
    default_name = "cv_processed.xlsx" if kind == "CV" else "lsv_processed.xlsx"
    out_path = filedialog.asksaveasfilename(
        title="Step 5 of 5: Save processed Excel file as...",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel files", "*.xlsx")],
        parent=root,
    )
    if not out_path:
        messagebox.showwarning("Cancelled", "No output location chosen.", parent=root)
        return False

    try:
        if kind == "CV":
            write_cv_workbook(df, table_df, scan_numbers, ph, ref_offset, area, out_path)
        else:
            write_lsv_workbook(df, ph, ref_offset, area, out_path)
    except Exception as e:
        messagebox.showerror(
            "Error saving file", f"{e}\n\nTip: if the file is currently open in Excel, close it first.",
            parent=root,
        )
        return False

    messagebox.showinfo(
        "Done!",
        f"Type: {kind}\nSaved: {out_path}\n\n"
        f"Reference: {ref_name} ({ref_offset} V)\npH: {ph}\nArea: {area} cm^2\n\n"
        "Open it in Excel -- pH / offset / area are editable in the yellow cells (AB2:AB4).",
        parent=root,
    )
    return True


def run_gui():
    root = tk.Tk()
    root.withdraw()

    while True:
        process_one_file(root)
        _bring_to_front(root)
        again = messagebox.askyesno(
            "Process another file?",
            "Do you want to process another raw data file (CV or LSV)?",
            parent=root,
        )
        if not again:
            break

    root.destroy()


if __name__ == "__main__":
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        script_dir = os.getcwd()
    log_path = os.path.join(script_dir, "cv_lsv_interactive_error.log")
    try:
        run_gui()
    except Exception:
        import traceback
        err = traceback.format_exc()
        print(err)
        with open(log_path, "w") as f:
            f.write(err)
        try:
            tk.Tk().withdraw()
            messagebox.showerror("Unexpected error", f"{err}\n\n(Also saved to {log_path})")
        except Exception:
            pass
        sys.exit(1)
