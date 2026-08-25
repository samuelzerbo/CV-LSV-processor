"""
app.py
========
CV / LSV Data Processor -- Streamlit web app.

Same working principle as the desktop cv_lsv_interactive.py tool, adapted
for the browser:

  0. Sign in with Google (st.login)
  1. Choose CV or LSV
  2. Upload raw data file -- validated against the required structure for
     the chosen type (same validation rules as the desktop tool, including
     the CV-vs-LSV 'Scan column' signature check)
  3. Pick reference electrode + enter pH
  4. Enter electrode surface area
  5. Process -> download the resulting Excel file

Run locally:
    pip install -r requirements.txt
    streamlit run app.py

Deploy: push this repo to GitHub, then deploy on
https://share.streamlit.io (Streamlit Community Cloud), free tier.
See README.md for full setup instructions (Google OAuth client, secrets).
"""

import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from gsheets_backend import backend_configured, log_login

# --------------------------- constants ---------------------------

SCAN_COL = "Scan"
INDEX_COL = "Index"
POTENTIAL_COL = "WE(1).Potential (V)"
CURRENT_COL = "WE(1).Current (A)"
REQUIRED_CV_SCANS = list(range(2, 11))

RAW_COLUMNS_ORDER_CV = [
    "Potential applied (V)", "Time (s)", "WE(1).Current (A)", "WE(1).Potential (V)",
    "Scan", "Index", "Q+", "Q-", "Current range",
]
RAW_COLUMNS_ORDER_LSV = [
    "Potential applied (V)", "Time (s)", "WE(1).Current (A)", "WE(1).Potential (V)",
    "Index", "Current range",
]

FONT = "Arial"
INPUT_FONT = Font(name=FONT, color="0000FF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
LABEL_FONT = Font(name=FONT, bold=True)
HEADER_FONT = Font(name=FONT, bold=True)
BODY_FONT = Font(name=FONT)

REF_ELECTRODES = [
    ("Ag/AgCl, saturated KCl", 0.197),
    ("Ag/AgCl, 3 M KCl", 0.210),
    ("Ag/AgCl, 3 M saturated KCl", 0.205),
    ("Ag/AgCl, 3 M NaCl", 0.209),
    ("Ag/AgCl, 1 M KCl", 0.235),
    ("SCE (saturated calomel)", 0.241),
    ("Hg/HgO, 1 M NaOH", 0.140),
]


# --------------------------- loading & validation ---------------------------

def load_raw_data(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded_file)
    else:
        df = pd.read_csv(uploaded_file, sep=None, engine="python")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def validate_structure(df: pd.DataFrame, kind: str):
    """Returns (ok: bool, message: str or None). Column-name validation plus
    the CV-vs-LSV 'Scan column' structural signature check."""
    if kind == "CV":
        required = [SCAN_COL, INDEX_COL, POTENTIAL_COL, CURRENT_COL]
    else:
        required = [INDEX_COL, POTENTIAL_COL, CURRENT_COL]

    missing = [c for c in required if c not in df.columns]
    if missing:
        return False, (
            f"This file is missing the following required column(s) for {kind} data: "
            f"**{', '.join(missing)}**.\n\n"
            f"Found columns: {', '.join(str(c) for c in df.columns)}\n\n"
            f"Please upload the correct raw {kind} data file."
        )

    if kind == "CV":
        scans_present = set(pd.to_numeric(df[SCAN_COL], errors="coerce").dropna().astype(int).unique().tolist())
        missing_scans = [s for s in REQUIRED_CV_SCANS if s not in scans_present]
        if missing_scans:
            return False, (
                f"This file's **{SCAN_COL}** column doesn't include all of scans 2 through 10.\n\n"
                f"Missing scan(s): {missing_scans}. Scans found: {sorted(scans_present)}\n\n"
                f"Please upload the correct raw CV data file (with scans 2-10)."
            )

    if kind == "LSV" and SCAN_COL in df.columns:
        scans_present = set(pd.to_numeric(df[SCAN_COL], errors="coerce").dropna().astype(int).unique().tolist())
        if len(scans_present) > 1:
            return False, (
                f"This file has a **{SCAN_COL}** column with {len(scans_present)} different scan values "
                f"({sorted(scans_present)}).\n\n"
                f"That's the signature of CV data (repeated cycles), not LSV (a single sweep).\n\n"
                f"Please choose **CV** instead, or upload the correct raw LSV data file."
            )

    return True, None


# --------------------------- CV processing ---------------------------

def build_cv_wide_table(df: pd.DataFrame, min_scan: int, max_scan: int, potential_from_scan: int):
    sub = df[df[SCAN_COL].between(min_scan, max_scan)].copy()
    sub = sub.sort_values([SCAN_COL, INDEX_COL])
    sub["_point"] = sub.groupby(SCAN_COL).cumcount()

    pivot = sub.pivot_table(index="_point", columns=SCAN_COL, values=CURRENT_COL)
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)

    pot_sub = sub[sub[SCAN_COL] == potential_from_scan][["_point", POTENTIAL_COL]]
    pot_sub = pot_sub.set_index("_point").rename(columns={POTENTIAL_COL: "Potential"})

    out = pot_sub.join(pivot, how="inner").reset_index(drop=True)
    scan_numbers = list(pivot.columns)
    return out, scan_numbers


def _write_inputs_block(ws, ph, ref_offset_v, area_cm2):
    in_col_label, in_col_value = 27, 28
    ws.cell(row=1, column=in_col_label, value="INPUTS (edit these)").font = LABEL_FONT

    ws.cell(row=2, column=in_col_label, value="pH").font = BODY_FONT
    c = ws.cell(row=2, column=in_col_value, value=ph); c.font, c.fill = INPUT_FONT, INPUT_FILL

    ws.cell(row=3, column=in_col_label, value="Reference electrode offset (V vs SHE)").font = BODY_FONT
    c = ws.cell(row=3, column=in_col_value, value=ref_offset_v); c.font, c.fill = INPUT_FONT, INPUT_FILL

    ws.cell(row=4, column=in_col_label, value="Electrode area (cm^2)").font = BODY_FONT
    c = ws.cell(row=4, column=in_col_value, value=area_cm2); c.font, c.fill = INPUT_FONT, INPUT_FILL

    ws.cell(row=5, column=in_col_label, value="Combined RHE offset (V) = offset + 0.0591*pH").font = BODY_FONT
    ws.cell(row=5, column=in_col_value,
            value=f"={get_column_letter(in_col_value)}3+0.0591*{get_column_letter(in_col_value)}2").font = BODY_FONT

    note = ws.cell(row=7, column=in_col_label,
                    value="pH, reference offset, and area are the values entered when this file was generated. "
                          "Edit the yellow cells above to change them -- the table recalculates automatically.")
    note.font = Font(name=FONT, italic=True, size=9)
    ws.merge_cells(start_row=7, start_column=in_col_label, end_row=9, end_column=in_col_value + 3)
    return in_col_label, in_col_value


def build_cv_excel_bytes(raw_df, table_df, scan_numbers, ph, ref_offset_v, area_cm2) -> bytes:
    wb = Workbook()
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, fullPrecision=True)
    ws = wb.active
    ws.title = "CV Data"

    raw_cols = [c for c in RAW_COLUMNS_ORDER_CV if c in raw_df.columns] or list(raw_df.columns)
    for j, col in enumerate(raw_cols, start=1):
        ws.cell(row=1, column=j, value=col).font = HEADER_FONT
    for i, row in enumerate(raw_df[raw_cols].itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val).font = BODY_FONT

    in_col_label, in_col_value = _write_inputs_block(ws, ph, ref_offset_v, area_cm2)

    start_col = 13
    headers = ["Potential"] + [str(s) for s in scan_numbers] + ["avarage", "Potential RHE", "current density mA/cm2"]
    for j, h in enumerate(headers, start=start_col):
        ws.cell(row=1, column=j, value=h).font = HEADER_FONT

    scan_col_letters = [get_column_letter(start_col + 1 + k) for k in range(len(scan_numbers))]
    pot_col_letter = get_column_letter(start_col)
    avg_col_letter = get_column_letter(start_col + 1 + len(scan_numbers))
    area_addr = f"${get_column_letter(in_col_value)}$4"
    combined_offset_addr = f"${get_column_letter(in_col_value)}$5"

    for i in range(len(table_df)):
        r = i + 2
        ws.cell(row=r, column=start_col, value=float(table_df["Potential"].iloc[i])).font = BODY_FONT
        for k, s in enumerate(scan_numbers):
            ws.cell(row=r, column=start_col + 1 + k, value=float(table_df[s].iloc[i])).font = BODY_FONT

        first_scan_col, last_scan_col = scan_col_letters[0], scan_col_letters[-1]
        ws.cell(row=r, column=start_col + 1 + len(scan_numbers),
                value=f"=AVERAGE({first_scan_col}{r}:{last_scan_col}{r})").font = BODY_FONT
        ws.cell(row=r, column=start_col + 2 + len(scan_numbers),
                value=f"={pot_col_letter}{r}+{combined_offset_addr}").font = BODY_FONT
        ws.cell(row=r, column=start_col + 3 + len(scan_numbers),
                value=f"=({avg_col_letter}{r}*1000)/{area_addr}").font = BODY_FONT

    for j in range(1, 30):
        ws.column_dimensions[get_column_letter(j)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------- LSV processing ---------------------------

def build_lsv_excel_bytes(raw_df, ph, ref_offset_v, area_cm2) -> bytes:
    wb = Workbook()
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, fullPrecision=True)
    ws = wb.active
    ws.title = "LSV Data"

    raw_cols = [c for c in RAW_COLUMNS_ORDER_LSV if c in raw_df.columns] or list(raw_df.columns)
    for j, col in enumerate(raw_cols, start=1):
        ws.cell(row=1, column=j, value=col).font = HEADER_FONT
    for i, row in enumerate(raw_df[raw_cols].itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val).font = BODY_FONT

    n_rows = len(raw_df)
    potential_col_letter = get_column_letter(raw_cols.index(POTENTIAL_COL) + 1)
    current_col_letter = get_column_letter(raw_cols.index(CURRENT_COL) + 1)

    in_col_label, in_col_value = _write_inputs_block(ws, ph, ref_offset_v, area_cm2)

    start_col = 13
    ws.cell(row=1, column=start_col, value="Potential RHE").font = HEADER_FONT
    ws.cell(row=1, column=start_col + 1, value="current density mA/cm2").font = HEADER_FONT

    area_addr = f"${get_column_letter(in_col_value)}$4"
    combined_offset_addr = f"${get_column_letter(in_col_value)}$5"

    for i in range(n_rows):
        r = i + 2
        ws.cell(row=r, column=start_col,
                value=f"={potential_col_letter}{r}+{combined_offset_addr}").font = BODY_FONT
        ws.cell(row=r, column=start_col + 1,
                value=f"=({current_col_letter}{r}*1000)/{area_addr}").font = BODY_FONT

    for j in range(1, 30):
        ws.column_dimensions[get_column_letter(j)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------- Streamlit UI ---------------------------

st.set_page_config(page_title="CV/LSV Data Processor", page_icon="🧪", layout="wide")

# ---- Background slideshow (blurred, crossfading CV <-> LSV plots) ----
import base64
import pathlib


def _img_b64(path):
    p = pathlib.Path(__file__).parent / path
    if not p.exists():
        return None
    return base64.b64encode(p.read_bytes()).decode()


_CV_B64 = _img_b64("cv_bg.jpg")
_LSV_B64 = _img_b64("lsv_bg.jpg")

if _CV_B64 and _LSV_B64:
    st.markdown(
        f"""
        <style>
        html, body {{
            background: transparent !important;
        }}
        [data-testid="stAppViewContainer"], [data-testid="stHeader"], [data-testid="stMain"], .stApp {{
            background: transparent !important;
        }}
        .bg-slide {{
            position: fixed;
            top: 0; left: 0;
            width: 100vw; height: 100vh;
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            filter: blur(9px);
            opacity: 0;
            z-index: -1;
            pointer-events: none;
            animation: bgFade 14s infinite;
        }}
        .bg-slide.cv {{
            background-image: url("data:image/jpeg;base64,{_CV_B64}");
            animation-delay: 0s;
        }}
        .bg-slide.lsv {{
            background-image: url("data:image/jpeg;base64,{_LSV_B64}");
            animation-delay: 7s;
        }}
        @keyframes bgFade {{
            0%   {{ opacity: 0; }}
            8%   {{ opacity: 0.30; }}
            42%  {{ opacity: 0.30; }}
            50%  {{ opacity: 0; }}
            100% {{ opacity: 0; }}
        }}
        </style>
        <div class="bg-slide cv"></div>
        <div class="bg-slide lsv"></div>
        """,
        unsafe_allow_html=True,
    )


def render_footer():
    st.markdown(
        "<div style='text-align:center; color:#999; font-size:12px; margin-top:32px;'>"
        "⚠️ ElectroProcess can make mistakes. Always verify the outputs."
        "</div>",
        unsafe_allow_html=True,
    )


# ---- Auth gate ----
try:
    logged_in = st.user.is_logged_in
except AttributeError:
    st.title("🧪 CV / LSV Data Processor")
    st.error(
        "Google sign-in isn't configured yet on this deployment.\n\n"
        "If you're the developer: add an `[auth]` block to `.streamlit/secrets.toml` "
        "(locally) or the app's Secrets panel (on Streamlit Community Cloud). "
        "See README.md for setup steps."
    )
    st.stop()

if not logged_in:
    left, mid, right = st.columns([1, 1.4, 1])
    with mid:
        st.markdown("<br><br>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown(
                "<div style='text-align:center; font-size:56px;'>🧪</div>",
                unsafe_allow_html=True,
            )
            st.markdown(
                "<h2 style='text-align:center; margin-top:0;'>CV / LSV Data Processor</h2>",
                unsafe_allow_html=True,
            )
            st.markdown(
                "<p style='text-align:center; color:#666;'>Convert raw voltammetry data to "
                "RHE potential &amp; current density in a few clicks.</p>",
                unsafe_allow_html=True,
            )
            st.write("")
            c1, c2, c3 = st.columns([1, 2, 1])
            with c2:
                if st.button("🔐  Log in with Google", type="primary", use_container_width=True):
                    st.login()
            st.markdown(
                "<div style='text-align:center; font-size:12px; color:#999;'>Your data is "
                "processed in memory for this session only.</div>",
                unsafe_allow_html=True,
            )
        render_footer()
    st.stop()

# ---- Log this login once per session (best-effort, never blocks the app) ----
if backend_configured() and not st.session_state.get("login_logged"):
    log_login(st.user.name, st.user.email)
    st.session_state.login_logged = True

# ---- TEMPORARY DEBUG BANNER: shows the real Sheets error, if any, right on
# the page. Remove this block once login logging is confirmed working. ----
if backend_configured() and st.session_state.get("_gsheets_last_error"):
    st.error(f"Google Sheets logging failed: {st.session_state['_gsheets_last_error']}")
elif not backend_configured():
    st.info("Debug: gcp_service_account/gsheets secrets not detected -- backend_configured() is False.")

# ---- Sidebar ----
with st.sidebar:
    st.markdown(
        f"""
        <div style="padding:14px; border-radius:10px; background:#F0F7F4; margin-bottom:14px;">
            <div style="font-size:13px; color:#666;">Signed in as</div>
            <div style="font-size:16px; font-weight:600;">{st.user.name}</div>
            <div style="font-size:12px; color:#888;">{st.user.email}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if st.button("Log out", use_container_width=True):
        st.logout()
    st.divider()
    st.caption("🧪 **CV / LSV Data Processor**")
    st.caption("Averages CV scans 2-10, or converts a single LSV sweep, to RHE potential and current density -- with live, editable formulas in the output.")

st.markdown(
    "<h1 style='margin-bottom:0;'>🧪 CV / LSV Data Processor</h1>"
    "<p style='color:#777; margin-top:4px;'>Upload &rarr; configure &rarr; download</p>",
    unsafe_allow_html=True,
)

# reset state if a fresh file is being processed
if "result_bytes" not in st.session_state:
    st.session_state.result_bytes = None
    st.session_state.result_name = None

if st.session_state.result_bytes is not None:
    with st.container(border=True):
        st.success("✅ Done! Your file is ready.")
        col1, col2 = st.columns([1, 1])
        with col1:
            st.download_button(
                "⬇️ Download processed Excel",
                data=st.session_state.result_bytes,
                file_name=st.session_state.result_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        with col2:
            if st.button("🔄 Process another file", use_container_width=True):
                st.session_state.result_bytes = None
                st.session_state.result_name = None
                st.rerun()
    render_footer()
    st.stop()

col_main, col_side = st.columns([2.3, 1])
uploaded, df, kind = None, None, None
ready = False  # becomes True once a valid file has been uploaded & validated

with col_main:
    with st.container(border=True):
        st.markdown("##### 1\ufe0f\u20e3 &nbsp; What type of data is this?")
        kind = st.radio(
            "Data type",
            ["CV", "LSV"],
            horizontal=True,
            captions=["Multiple scans, averages scans 2-10", "Single sweep"],
            label_visibility="collapsed",
        )

    with st.container(border=True):
        st.markdown("##### 2\ufe0f\u20e3 &nbsp; Upload your raw data file")
        uploaded = st.file_uploader("Raw data file", type=["csv", "xlsx", "xls"], label_visibility="collapsed")

        if uploaded is None:
            st.info("Waiting for a file...")
        else:
            try:
                df = load_raw_data(uploaded)
            except Exception as e:
                st.error(f"Could not read this file: {e}")
                df = None

            if df is not None:
                ok, msg = validate_structure(df, kind)
                if not ok:
                    st.error(msg)
                else:
                    st.success(f"'{uploaded.name}' loaded ({len(df):,} rows) -- looks like valid {kind} data.")
                    ready = True

    if ready:
        with st.container(border=True):
            st.markdown("##### 3\ufe0f\u20e3 &nbsp; Reference electrode & pH")
            ref_labels = [f"{name} (offset = {offset} V)" for name, offset in REF_ELECTRODES] + ["Custom offset..."]
            choice = st.selectbox("Reference electrode", ref_labels)
            if choice == "Custom offset...":
                ref_name = "Custom"
                ref_offset = st.number_input("Custom offset (V vs SHE)", value=0.000, format="%.4f")
            else:
                ref_name, ref_offset = REF_ELECTRODES[ref_labels.index(choice)]

            ph = st.number_input("Electrolyte pH", min_value=0.0, max_value=14.0, value=7.0, step=0.1)
            st.caption(f"📐 E_RHE = E_measured + {ref_offset} + 0.0591 × pH")

        with st.container(border=True):
            st.markdown("##### 4\ufe0f\u20e3 &nbsp; Electrode surface area")
            area = st.number_input("Electrode surface area (cm²)", min_value=0.000001, value=0.070, format="%.4f")

        with st.container(border=True):
            st.markdown("##### 5\ufe0f\u20e3 &nbsp; Process")
            if st.button("⚙️  Process file", type="primary", use_container_width=True):
                with st.spinner("Processing..."):
                    try:
                        if kind == "CV":
                            table_df, scan_numbers = build_cv_wide_table(df, min_scan=2, max_scan=10, potential_from_scan=2)
                            out_bytes = build_cv_excel_bytes(df, table_df, scan_numbers, ph, ref_offset, area)
                            out_name = "cv_processed.xlsx"
                        else:
                            out_bytes = build_lsv_excel_bytes(df, ph, ref_offset, area)
                            out_name = "lsv_processed.xlsx"
                    except Exception as e:
                        st.error(f"Error processing data: {e}")
                        out_bytes = None

                if out_bytes is not None:
                    st.session_state.result_bytes = out_bytes
                    st.session_state.result_name = out_name
                    st.rerun()

with col_side:
    st.markdown("##### 📄 File summary")
    if uploaded is not None and df is not None:
        with st.container(border=True):
            st.metric("Rows", f"{len(df):,}")
            st.metric("Data type", kind)
            if kind == "CV" and SCAN_COL in df.columns:
                st.metric("Scans found", df[SCAN_COL].nunique())
    st.markdown("##### ⚗️ Conversion formulas")
    with st.container(border=True):
        st.caption("**RHE potential**")
        st.markdown("`E_RHE = E + offset + 0.0591·pH`")
        st.caption("**Current density**")
        st.markdown("`j = (I × 1000) / area`")

render_footer()
